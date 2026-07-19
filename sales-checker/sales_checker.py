# -*- coding: utf-8 -*-
"""
매출 확인 프로그램 (Cafe24 / 더망고 주문 엑셀 기반)
----------------------------------------------------
- 지정한 쇼핑몰 번호(377 / 2552 / 2179)의 주문 엑셀을 자동으로 내려받아
  아래 4가지를 정리해 브라우저 대시보드로 보여줍니다.

  1) 엑셀 최상단의 '마지막주문수집일자'
  2) '더망고주문상태' 중 "반품/교환/취소 진행중", "반품/교환/취소완료"를
     제외한 나머지의 '총 결제금액합계(원)' -> 해당 월 총 매출
  3) '마켓주문일자'별 매출(위 상태 제외)
  4) '마켓명'별 매출(위 상태 제외)

- 별도 서버/웹페이지 배포 없이, 실행 파일을 실행하면 로컬에서 페이지가 열립니다.
"""

import io
import os
import re
import sys
import json
import ssl
import socket
import threading
import webbrowser
from datetime import datetime, timezone, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlencode, quote, parse_qs, urlparse
from urllib.request import Request, urlopen

# ----------------------------------------------------------------------------
# 설정
# ----------------------------------------------------------------------------

# 확인할 쇼핑몰 (번호 = tmg<번호>.cafe24.com). 이름은 보기 좋게 바꿔도 됩니다.
STORES = [
    {"code": "377",  "name": "어현준"},
    {"code": "2552", "name": "이동욱"},
    {"code": "2179", "name": "황민호"},
]

# 제외할 더망고 주문상태
EXCLUDE_STATUS = {"반품/교환/취소 진행중", "반품/교환/취소완료"}

# 카테고리별 매출 집계 기간(일). 최근 N일(오늘 포함) 데이터만 집계.
CATEGORY_DAYS = 3

# '롯데ON' 마켓 판별용 키워드(공백/대소문자 무시)
LOTTEON_KEYS = ["롯데on", "롯데온", "lotteon"]

# 엑셀 컬럼 매칭 후보 (부분 일치, 앞에 있을수록 우선)
COL_CANDIDATES = {
    "status": ["더망고주문상태", "주문상태"],
    "amount": ["총 결제금액합계(원)", "총결제금액합계", "결제금액합계", "결제금액", "실결제금액"],
    "order_date": ["마켓주문일자", "주문일자", "주문일시", "결제일"],
    "market": ["마켓명", "마켓", "판매처"],
    "category": ["카테고리", "상품카테고리", "카테고리명", "상품분류", "분류", "상품분류명", "진열카테고리"],
}

KST = timezone(timedelta(hours=9))
HOST = "127.0.0.1"
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")


# ----------------------------------------------------------------------------
# 기간(월) -> Cafe24 search_sql 다운로드 URL
# ----------------------------------------------------------------------------

def month_range_epoch(year: int, month: int):
    """해당 월의 시작/끝 UNIX 타임스탬프(KST 기준)."""
    start = datetime(year, month, 1, tzinfo=KST)
    if month == 12:
        nxt = datetime(year + 1, 1, 1, tzinfo=KST)
    else:
        nxt = datetime(year, month + 1, 1, tzinfo=KST)
    return int(start.timestamp()), int(nxt.timestamp()) - 1


def build_download_url(code: str, year: int, month: int) -> str:
    start, end = month_range_epoch(year, month)
    search_sql = (
        " and A.register_date >= '%d' and A.register_date <= '%d' "
        " and ((A.buyer_payment = '1' and A.buyer_state > 1) or A.buyer_payment != '1')"
        % (start, end)
    )
    qs = urlencode(
        {"uid_check": "", "excel_type": "all", "search_sql": search_sql},
        quote_via=quote,
    )
    return "https://tmg%s.cafe24.com/mall/admin/admin_excel2.php?%s" % (code, qs)


# ----------------------------------------------------------------------------
# 다운로드
# ----------------------------------------------------------------------------

def download_bytes(url: str, timeout: int = 60) -> bytes:
    ctx = ssl.create_default_context()
    req = Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) SalesChecker/1.0",
        "Accept": "*/*",
    })
    with urlopen(req, timeout=timeout, context=ctx) as resp:
        return resp.read()


def local_file_for(code: str):
    """data 폴더에 수동으로 넣어둔 파일(예: 377.xls)이 있으면 경로 반환."""
    if not os.path.isdir(DATA_DIR):
        return None
    for fn in os.listdir(DATA_DIR):
        base, ext = os.path.splitext(fn)
        if base == code and ext.lower() in (".xls", ".xlsx", ".htm", ".html", ".csv"):
            return os.path.join(DATA_DIR, fn)
    return None


# ----------------------------------------------------------------------------
# 엑셀 -> 2차원 그리드(문자열)  (포맷 자동 판별)
# ----------------------------------------------------------------------------

class _TableGridParser:
    """HTML 표를 그리드로 변환하는 최소 파서."""
    def __init__(self):
        from html.parser import HTMLParser

        rows = []
        cur = []
        buf = []
        state = {"in_cell": False}

        class P(HTMLParser):
            def handle_starttag(self, tag, attrs):
                if tag in ("td", "th"):
                    state["in_cell"] = True
                    buf.clear()
                elif tag == "br" and state["in_cell"]:
                    buf.append(" ")

            def handle_endtag(self, tag):
                if tag in ("td", "th"):
                    state["in_cell"] = False
                    cur.append("".join(buf).strip())
                    buf.clear()
                elif tag == "tr":
                    if cur:
                        rows.append(list(cur))
                    cur.clear()

            def handle_data(self, data):
                if state["in_cell"]:
                    buf.append(data)

        self._parser = P()
        self._rows = rows

    def feed(self, text):
        self._parser.feed(text)
        return self._rows


def _decode_bytes(raw: bytes) -> str:
    for enc in ("cp949", "euc-kr", "utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(enc)
        except Exception:
            continue
    return raw.decode("utf-8", errors="replace")


def load_grid(raw: bytes):
    """엑셀/HTML/CSV 바이트를 문자열 2차원 배열로 변환."""
    if not raw:
        return []

    head = raw[:8]

    # 1) xlsx (zip)
    if head[:4] == b"PK\x03\x04":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        grid = []
        for row in ws.iter_rows(values_only=True):
            grid.append(["" if c is None else str(c) for c in row])
        wb.close()
        return grid

    # 2) 구형 xls (OLE)
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        import xlrd
        book = xlrd.open_workbook(file_contents=raw)
        sh = book.sheet_by_index(0)
        grid = []
        for r in range(sh.nrows):
            grid.append([_xls_cell_str(sh.cell(r, c)) for c in range(sh.ncols)])
        return grid

    # 3) HTML / CSV / TSV (Cafe24 admin_excel2.php 는 보통 EUC-KR HTML table)
    text = _decode_bytes(raw)
    lower = text.lower()
    if "<table" in lower or "<td" in lower or "<tr" in lower:
        grid = _TableGridParser().feed(text)
        if grid:
            return grid

    # CSV / TSV fallback
    lines = [ln for ln in text.splitlines() if ln.strip() != ""]
    delim = "\t" if (lines and lines[0].count("\t") >= lines[0].count(",")) else ","
    return [ln.split(delim) for ln in lines]


def _xls_cell_str(cell):
    import xlrd
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        v = cell.value
        return str(int(v)) if float(v).is_integer() else str(v)
    return "" if cell.value is None else str(cell.value)


# ----------------------------------------------------------------------------
# 분석
# ----------------------------------------------------------------------------

def _norm(s: str) -> str:
    return re.sub(r"\s+", "", str(s or ""))


def _to_amount(s) -> int:
    if s is None:
        return 0
    t = str(s).replace(",", "").strip()
    m = re.search(r"-?\d+", t)
    return int(m.group()) if m else 0


def _find_col(headers, keys):
    hn = [_norm(h) for h in headers]
    for key in keys:
        nk = _norm(key)
        for i, h in enumerate(hn):
            if h == nk:
                return i
    for key in keys:
        nk = _norm(key)
        for i, h in enumerate(hn):
            if nk and nk in h:
                return i
    return -1


def _find_last_collect_date(grid, header_row_idx):
    """'마지막주문수집일자' 라벨을 찾아 값을 반환."""
    limit = header_row_idx if header_row_idx >= 0 else min(len(grid), 15)
    for r in range(min(limit + 1, len(grid))):
        for c, cell in enumerate(grid[r]):
            if "마지막주문수집일자" in _norm(cell):
                # 같은 셀 안에 값이 있는 경우 (예: "마지막주문수집일자 : 2026-07-19 ...")
                m = re.search(r"마지막주문수집일자\s*[:：]?\s*(.+)", str(cell))
                if m and m.group(1).strip():
                    return m.group(1).strip()
                # 오른쪽 셀에 값이 있는 경우
                for c2 in range(c + 1, len(grid[r])):
                    if str(grid[r][c2]).strip():
                        return str(grid[r][c2]).strip()
    return ""


def _find_header_row(grid):
    """알려진 컬럼명이 가장 많이 등장하는 행을 헤더로 판단."""
    all_keys = [k for keys in COL_CANDIDATES.values() for k in keys]
    best_idx, best_score = -1, 0
    for r in range(min(len(grid), 30)):
        row_norm = [_norm(c) for c in grid[r]]
        score = 0
        for key in all_keys:
            nk = _norm(key)
            if any(nk and nk in cell for cell in row_norm):
                score += 1
        if score > best_score:
            best_score, best_idx = score, r
    return best_idx if best_score >= 2 else -1


def _parse_ymd(s):
    """'2026-07-15 13:22' / '2026.07.15' / '20260715' -> ('2026-07-15','2026-07')"""
    t = str(s or "")
    m = re.search(r"(\d{4})[.\-/]?\s?(\d{1,2})[.\-/]?\s?(\d{1,2})", t)
    if not m:
        return "", ""
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return "%04d-%02d-%02d" % (y, mo, d), "%04d-%02d" % (y, mo)


def _parse_dt(s):
    """마켓주문일자 셀 -> (datetime, 표시문자열). 파싱 실패 시 (None, '')."""
    t = str(s or "")
    m = re.search(
        r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})(?:[ T]+(\d{1,2}):(\d{2})(?::(\d{2}))?)?",
        t,
    )
    if not m:
        return None, ""
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    hh = int(m.group(4)) if m.group(4) else 0
    mi = int(m.group(5)) if m.group(5) else 0
    ss = int(m.group(6)) if m.group(6) else 0
    try:
        dt = datetime(y, mo, d, hh, mi, ss, tzinfo=KST)
    except ValueError:
        return None, ""
    if m.group(4):
        disp = "%04d-%02d-%02d %02d:%02d:%02d" % (y, mo, d, hh, mi, ss)
    else:
        disp = "%04d-%02d-%02d" % (y, mo, d)
    return dt, disp


def _is_lotteon(market):
    n = _norm(market).lower()
    return any(k in n for k in LOTTEON_KEYS)


def analyze_grid(grid, target_ym):
    now = datetime.now(KST)
    today_str = now.strftime("%Y-%m-%d")
    cat_from = (now - timedelta(days=CATEGORY_DAYS - 1)).strftime("%Y-%m-%d")
    result = {
        "last_collect_date": "",
        "last_order_date": "",
        "target_ym": target_ym,
        "month_total": 0,
        "today": today_str,
        "today_total": 0,
        "cat_days": CATEGORY_DAYS,
        "cat_from": cat_from,
        "cat_to": today_str,
        "row_count": 0,
        "excluded_count": 0,
        "by_date": [],
        "lotteon_by_date": [],
        "by_market": [],
        "by_category": [],
        "warnings": [],
    }
    if not grid:
        result["warnings"].append("파일이 비어있거나 읽을 수 없습니다.")
        return result

    hidx = _find_header_row(grid)
    result["last_collect_date"] = _find_last_collect_date(grid, hidx)

    if hidx < 0:
        result["warnings"].append("엑셀에서 표 머리글(마켓주문일자/총 결제금액합계 등)을 찾지 못했습니다.")
        return result

    headers = grid[hidx]
    ci = {k: _find_col(headers, v) for k, v in COL_CANDIDATES.items()}
    if ci["amount"] < 0:
        result["warnings"].append("'총 결제금액합계(원)' 컬럼을 찾지 못했습니다.")
        return result

    by_date, by_market, by_month, by_category = {}, {}, {}, {}
    lotteon_by_date = {}
    total_rows = excluded = 0
    best_dt, best_disp = None, ""       # 미래 제외, 현재에 가장 근접한 최신 주문일시
    max_dt, max_disp = None, ""         # 전체 최신(폴백)
    cat_from_date = (now - timedelta(days=CATEGORY_DAYS - 1)).date()

    for r in range(hidx + 1, len(grid)):
        row = grid[r]
        if ci["amount"] >= len(row):
            continue
        amount_cell = row[ci["amount"]]
        status = row[ci["status"]] if 0 <= ci["status"] < len(row) else ""
        # 합계/빈 행 건너뛰기
        if _to_amount(amount_cell) == 0 and not str(amount_cell).strip().lstrip("-").isdigit():
            if not any(str(x).strip() for x in row):
                continue
        total_rows += 1

        if _norm(status) in {_norm(s) for s in EXCLUDE_STATUS}:
            excluded += 1
            continue

        amt = _to_amount(amount_cell)
        odate = row[ci["order_date"]] if 0 <= ci["order_date"] < len(row) else ""
        market = row[ci["market"]] if 0 <= ci["market"] < len(row) else ""
        market = str(market).strip() or "(마켓명 없음)"
        category = row[ci["category"]] if 0 <= ci["category"] < len(row) else ""
        category = str(category).strip() or "(카테고리 없음)"
        ymd, ym = _parse_ymd(odate)
        odt, odisp = _parse_dt(odate)

        # 마지막 마켓주문일자 계산(마켓주문일자 열 기준)
        if odt is not None:
            if max_dt is None or odt > max_dt:
                max_dt, max_disp = odt, odisp
            if odt <= now and (best_dt is None or odt > best_dt):
                best_dt, best_disp = odt, odisp

        if ym:
            by_month[ym] = by_month.get(ym, 0) + amt
        if ymd:
            by_date[ymd] = by_date.get(ymd, 0) + amt
            if _is_lotteon(market):
                lotteon_by_date[ymd] = lotteon_by_date.get(ymd, 0) + amt
        by_market[market] = by_market.get(market, 0) + amt
        # 카테고리별 매출은 최근 CATEGORY_DAYS일(오늘 포함) 데이터만 집계
        if ci["category"] >= 0 and odt is not None and odt.date() >= cat_from_date:
            c = by_category.setdefault(category, {"total": 0, "count": 0})
            c["total"] += amt
            c["count"] += 1

    result["row_count"] = total_rows
    result["excluded_count"] = excluded
    result["month_total"] = by_month.get(target_ym, 0)
    result["today_total"] = by_date.get(today_str, 0)
    result["last_order_date"] = best_disp or max_disp
    result["all_months"] = [{"ym": k, "total": v} for k, v in sorted(by_month.items())]
    result["by_date"] = [
        {"date": k, "total": v} for k, v in sorted(by_date.items())
    ]
    result["lotteon_by_date"] = [
        {"date": k, "total": v} for k, v in sorted(lotteon_by_date.items())
    ]
    result["by_market"] = [
        {"market": k, "total": v} for k, v in sorted(by_market.items(), key=lambda x: -x[1])
    ]
    result["by_category"] = [
        {"category": k, "total": v["total"], "count": v["count"]}
        for k, v in sorted(by_category.items(), key=lambda x: -x[1]["total"])
    ]
    if ci["category"] < 0:
        result["warnings"].append("엑셀에서 '카테고리' 컬럼을 찾지 못했습니다.")
    return result


def analyze_store(store, year, month):
    ym = "%04d-%02d" % (year, month)
    out = {"code": store["code"], "name": store["name"], "ok": False, "source": "",
           "error": "", "url": build_download_url(store["code"], year, month)}
    raw = None
    try:
        raw = download_bytes(out["url"])
        out["source"] = "다운로드"
    except Exception as e:
        # 다운로드 실패 시 로컬 파일 폴백
        lf = local_file_for(store["code"])
        if lf:
            try:
                with open(lf, "rb") as f:
                    raw = f.read()
                out["source"] = "로컬파일(%s)" % os.path.basename(lf)
            except Exception as e2:
                out["error"] = "로컬 파일 읽기 실패: %s" % e2
        else:
            out["error"] = "다운로드 실패: %s" % e
    if raw is None:
        return out
    try:
        grid = load_grid(raw)
        data = analyze_grid(grid, ym)
        out.update(data)
        out["ok"] = True
    except Exception as e:
        out["error"] = "분석 실패: %s" % e
    return out


def analyze_all(year, month):
    results = []
    threads = []
    slots = [None] * len(STORES)

    def work(i, st):
        slots[i] = analyze_store(st, year, month)

    for i, st in enumerate(STORES):
        t = threading.Thread(target=work, args=(i, st))
        t.start()
        threads.append(t)
    for t in threads:
        t.join()
    results = [s for s in slots if s]
    return {
        "year": year, "month": month, "ym": "%04d-%02d" % (year, month),
        "generated_at": datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S"),
        "stores": results,
    }


# ----------------------------------------------------------------------------
# 로컬 웹 서버 + UI
# ----------------------------------------------------------------------------

PAGE_HTML = r"""<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>매출 확인</title>
<style>
  :root{
    --bg:#0f1220; --card:#171a2b; --line:#2a2f45; --txt:#e8eaf2; --sub:#9aa0be;
    --accent:#5b8cff; --good:#37d39b; --warn:#ffb020; --bad:#ff6b6b;
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--bg);color:var(--txt);
    font-family:"Malgun Gothic","맑은 고딕",-apple-system,Segoe UI,Roboto,sans-serif;}
  header{position:sticky;top:0;z-index:5;background:rgba(15,18,32,.92);
    backdrop-filter:blur(6px);border-bottom:1px solid var(--line);padding:14px 20px;}
  .bar{display:flex;align-items:center;gap:12px;flex-wrap:wrap;max-width:1200px;margin:0 auto;}
  h1{font-size:18px;margin:0;font-weight:700}
  .grow{flex:1}
  select,button{font:inherit;color:var(--txt);background:#21263d;border:1px solid var(--line);
    border-radius:10px;padding:9px 14px;cursor:pointer}
  button.primary{background:var(--accent);border-color:var(--accent);color:#fff;font-weight:700}
  button:disabled{opacity:.5;cursor:default}
  .meta{color:var(--sub);font-size:12px}
  main{max-width:1200px;margin:0 auto;padding:20px}
  .grand{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:18px}
  .grand .box{flex:1;min-width:220px;background:var(--card);
    border:1px solid var(--line);border-radius:16px;padding:18px 20px}
  .grand .box.today{background:linear-gradient(160deg,#1d3352,#182034);border-color:var(--accent)}
  .grand .box.today .val{color:#7ea9ff}
  .grand .label{color:var(--sub);font-size:13px}
  .grand .val{font-size:26px;font-weight:800;margin-top:6px;letter-spacing:-.5px}
  .grand .sub{color:var(--sub);font-size:12px;margin-top:6px}
  .stores{display:grid;grid-template-columns:1fr;gap:18px}
  @media(min-width:900px){.stores{grid-template-columns:1fr 1fr 1fr}}
  .card{background:var(--card);border:1px solid var(--line);border-radius:16px;overflow:hidden;
    display:flex;flex-direction:column}
  .card h2{margin:0;padding:16px 18px;font-size:16px;border-bottom:1px solid var(--line);
    display:flex;align-items:center;gap:8px}
  .pill{font-size:11px;padding:3px 8px;border-radius:999px;background:#232842;color:var(--sub)}
  .card .body{padding:16px 18px;display:flex;flex-direction:column;gap:16px}
  .kv{display:flex;justify-content:space-between;gap:10px;font-size:13px}
  .kv .k{color:var(--sub)}
  .big{font-size:22px;font-weight:800}
  .sec-title{font-size:12px;color:var(--sub);text-transform:uppercase;letter-spacing:.4px;
    margin-bottom:6px;font-weight:700}
  table{width:100%;border-collapse:collapse;font-size:13px}
  th,td{padding:7px 8px;text-align:left;border-bottom:1px solid var(--line)}
  th{color:var(--sub);font-weight:600}
  td.num,th.num{text-align:right;font-variant-numeric:tabular-nums}
  .scroll{max-height:230px;overflow:auto;border:1px solid var(--line);border-radius:10px}
  .scroll table th{position:sticky;top:0;background:#1b2038}
  .err{color:var(--bad);font-size:13px;line-height:1.5}
  .warn{color:var(--warn);font-size:12px;margin-top:4px}
  .foot{color:var(--sub);font-size:11px;margin-top:8px;word-break:break-all}
  .loading{display:flex;align-items:center;gap:10px;color:var(--sub);padding:40px;justify-content:center}
  .spin{width:16px;height:16px;border:2px solid var(--line);border-top-color:var(--accent);
    border-radius:50%;animation:sp 1s linear infinite}
  @keyframes sp{to{transform:rotate(360deg)}}
</style>
</head>
<body>
<header>
  <div class="bar">
    <h1>📊 매출 확인</h1>
    <select id="month"></select>
    <button class="primary" id="run">매출 확인</button>
    <div class="grow"></div>
    <div class="meta" id="stamp"></div>
  </div>
</header>
<main>
  <div class="grand" id="grand"></div>
  <div id="content">
    <div class="loading"><div class="spin"></div> 준비되었습니다. 상단의 <b>&nbsp;매출 확인&nbsp;</b> 을 눌러주세요.</div>
  </div>
</main>
<script>
const won = n => (n<0?"-":"") + Math.abs(n).toLocaleString("ko-KR") + "원";
function fillMonths(){
  const sel = document.getElementById("month");
  const now = new Date();
  for(let i=0;i<18;i++){
    const d = new Date(now.getFullYear(), now.getMonth()-i, 1);
    const ym = d.getFullYear()+"-"+String(d.getMonth()+1).padStart(2,"0");
    const o = document.createElement("option");
    o.value = ym; o.textContent = d.getFullYear()+"년 "+(d.getMonth()+1)+"월";
    sel.appendChild(o);
  }
}
function storeCard(s){
  if(!s.ok){
    return `<div class="card"><h2>${s.name} <span class="pill">쇼핑몰 ${s.code}</span></h2>
      <div class="body"><div class="err">⚠️ ${s.error||"불러오지 못했습니다."}</div>
      <div class="foot">요청 주소: ${s.url}</div></div></div>`;
  }
  const byDate = (s.by_date||[]).map(r=>`<tr><td>${r.date}</td><td class="num">${won(r.total)}</td></tr>`).join("") || `<tr><td colspan="2">데이터 없음</td></tr>`;
  const lotte = (s.lotteon_by_date||[]).map(r=>`<tr><td>${r.date}</td><td class="num">${won(r.total)}</td></tr>`).join("") || `<tr><td colspan="2">롯데ON 매출 없음</td></tr>`;
  const byMk = (s.by_market||[]).map(r=>`<tr><td>${r.market}</td><td class="num">${won(r.total)}</td></tr>`).join("") || `<tr><td colspan="2">데이터 없음</td></tr>`;
  const byCat = (s.by_category||[]).map(r=>`<tr><td>${r.category}</td><td class="num">${(r.count||0).toLocaleString("ko-KR")}건</td><td class="num">${won(r.total)}</td></tr>`).join("") || `<tr><td colspan="3">최근 ${s.cat_days||3}일 카테고리 데이터 없음</td></tr>`;
  const warn = (s.warnings&&s.warnings.length)?`<div class="warn">※ ${s.warnings.join(" / ")}</div>`:"";
  return `<div class="card">
    <h2>${s.name} <span class="pill">쇼핑몰 ${s.code}</span> <span class="pill">${s.source||""}</span></h2>
    <div class="body">
      <div>
        <div class="kv"><span class="k">마지막마켓주문일자</span><span>${s.last_order_date||"-"}</span></div>
      </div>
      <div>
        <div class="sec-title">${s.target_ym} 총 매출 (반품/교환/취소 제외)</div>
        <div class="big">${won(s.month_total||0)}</div>
        <div class="kv" style="margin-top:6px"><span class="k">오늘(${s.today||""}) 매출</span><span>${won(s.today_total||0)}</span></div>
        <div class="kv" style="margin-top:4px"><span class="k">집계 건수</span><span>${(s.row_count-s.excluded_count).toLocaleString("ko-KR")}건 (제외 ${s.excluded_count}건)</span></div>
      </div>
      <div>
        <div class="sec-title">롯데ON 일자별 매출</div>
        <div class="scroll"><table><thead><tr><th>일자</th><th class="num">매출</th></tr></thead><tbody>${lotte}</tbody></table></div>
      </div>
      <div>
        <div class="sec-title">마켓주문일자별 매출</div>
        <div class="scroll"><table><thead><tr><th>일자</th><th class="num">매출</th></tr></thead><tbody>${byDate}</tbody></table></div>
      </div>
      <div>
        <div class="sec-title">마켓명별 매출</div>
        <div class="scroll"><table><thead><tr><th>마켓명</th><th class="num">매출</th></tr></thead><tbody>${byMk}</tbody></table></div>
      </div>
      <div>
        <div class="sec-title">상품 카테고리별 매출 <span style="font-weight:400;text-transform:none">(최근 ${s.cat_days||3}일: ${s.cat_from||""} ~ ${s.cat_to||""})</span></div>
        <div class="scroll"><table><thead><tr><th>카테고리</th><th class="num">건수</th><th class="num">매출</th></tr></thead><tbody>${byCat}</tbody></table></div>
      </div>
      ${warn}
    </div>
  </div>`;
}
async function run(){
  const btn = document.getElementById("run");
  const ym = document.getElementById("month").value;
  btn.disabled = true; btn.textContent = "확인 중...";
  document.getElementById("content").innerHTML =
    '<div class="loading"><div class="spin"></div> 3개 쇼핑몰의 엑셀을 내려받아 분석 중입니다...</div>';
  document.getElementById("grand").innerHTML = "";
  try{
    const res = await fetch("/api/data?ym="+encodeURIComponent(ym));
    const data = await res.json();
    render(data);
  }catch(e){
    document.getElementById("content").innerHTML = '<div class="err">오류: '+e+'</div>';
  }finally{
    btn.disabled = false; btn.textContent = "매출 확인";
  }
}
function render(data){
  document.getElementById("stamp").textContent = "생성 "+data.generated_at;
  const oks = data.stores.filter(s=>s.ok);
  const grandMonth = oks.reduce((a,s)=>a+(s.month_total||0),0);
  const grandToday = oks.reduce((a,s)=>a+(s.today_total||0),0);
  const today = (oks[0] && oks[0].today) || "";
  const orderDates = oks.map(s=>s.last_order_date).filter(Boolean);
  const lastOrder = orderDates.length ? orderDates.sort().slice(-1)[0] : "-";
  document.getElementById("grand").innerHTML =
    `<div class="box today"><div class="label">오늘 (${today}) 3개 쇼핑몰 합계 매출</div>`
    + `<div class="val">${won(grandToday)}</div>`
    + `<div class="sub">${data.ym} 이달 합계 ${won(grandMonth)}</div>`
    + `<div class="sub">마지막마켓주문일자 ${lastOrder}</div></div>`
    + data.stores.map(s=>`<div class="box"><div class="label">${s.name} (${s.code})</div>`
        + `<div class="val">${s.ok?won(s.today_total||0):"-"}</div>`
        + `<div class="sub">${s.ok?("이달 "+won(s.month_total||0)):"불러오기 실패"}</div>`
        + `<div class="sub">${s.ok?("마지막주문 "+(s.last_order_date||"-")):""}</div></div>`).join("");
  document.getElementById("content").innerHTML =
    '<div class="stores">'+data.stores.map(storeCard).join("")+'</div>';
}
fillMonths();
document.getElementById("run").addEventListener("click", run);
// 페이지 로드시 자동 실행
window.addEventListener("load", run);
</script>
</body>
</html>"""


class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a):
        pass

    def _send(self, code, ctype, body):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path in ("/", "/index.html"):
            self._send(200, "text/html; charset=utf-8", PAGE_HTML.encode("utf-8"))
            return
        if parsed.path == "/api/data":
            q = parse_qs(parsed.query)
            ym = (q.get("ym") or [""])[0]
            try:
                year, month = map(int, ym.split("-"))
            except Exception:
                now = datetime.now(KST)
                year, month = now.year, now.month
            try:
                data = analyze_all(year, month)
                body = json.dumps(data, ensure_ascii=False).encode("utf-8")
                self._send(200, "application/json; charset=utf-8", body)
            except Exception as e:
                body = json.dumps({"error": str(e)}, ensure_ascii=False).encode("utf-8")
                self._send(500, "application/json; charset=utf-8", body)
            return
        self._send(404, "text/plain; charset=utf-8", b"not found")


def find_free_port():
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind((HOST, 0))
    port = s.getsockname()[1]
    s.close()
    return port


def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    port = find_free_port()
    httpd = ThreadingHTTPServer((HOST, port), Handler)
    url = "http://%s:%d/" % (HOST, port)
    print("=" * 52)
    print(" 매출 확인 프로그램이 실행되었습니다.")
    print(" 브라우저에서 다음 주소가 열립니다:")
    print("   " + url)
    print(" 이 창을 닫으면 프로그램이 종료됩니다.")
    print("=" * 52)
    try:
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    except Exception:
        pass
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n종료합니다.")
        httpd.shutdown()


if __name__ == "__main__":
    main()
